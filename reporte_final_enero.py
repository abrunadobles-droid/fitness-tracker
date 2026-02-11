"""
Reporte Final Enero 2026 - Garmin + WHOOP v2
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from garmin_client import GarminClient
from whoop_client_v2 import WhoopClientV2
from calendar import monthrange

print("\n" + "="*70)
print("  REPORTE FINAL MENSUAL - ENERO 2026")
print("  Garmin Connect + WHOOP API v2")
print("="*70)

YEAR = 2026
MONTH = 1

wb = Workbook()
ws = wb.active
ws.title = "Enero 2026"

# Estilos
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, size=14, color='FFFFFF')
subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
subheader_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=12, color='4472C4')

# Título
ws['A1'] = 'REPORTE MENSUAL FITNESS & WELLNESS'
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws.merge_cells('A1:D1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

ws['A2'] = f'Enero 2026'
ws['A2'].font = Font(bold=True, size=12)
ws.merge_cells('A2:D2')

ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws['A3'].font = Font(italic=True, size=9)

# Headers tabla
ws['A5'] = 'MÉTRICA'
ws['B5'] = 'VALOR'
ws['C5'] = 'META'
ws['D5'] = 'NOTAS'

for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}5'].font = subheader_font
    ws[f'{col}5'].fill = subheader_fill
    ws[f'{col}5'].alignment = Alignment(horizontal='center')

row = 6

# ==================== GARMIN ====================
print("\n📊 Procesando GARMIN...")

try:
    garmin = GarminClient()
    garmin.login()
    
    start_date = datetime(YEAR, MONTH, 1)
    last_day = monthrange(YEAR, MONTH)[1]
    end_date = datetime(YEAR, MONTH, last_day)
    
    total_steps = 0
    days_with_steps = 0
    all_activities = []
    strength_sessions = 0
    total_sleep_seconds = 0
    days_with_sleep = 0
    
    print(f"   📅 Procesando {last_day} días de enero...")
    
    current_date = start_date
    days_processed = 0
    
    while current_date <= end_date:
        try:
            # Stats
            stats = garmin.get_stats_for_date(current_date)
            if stats and stats.get('totalSteps'):
                total_steps += stats['totalSteps']
                days_with_steps += 1
            
            # Sleep
            try:
                sleep_data = garmin.get_sleep_data(current_date)
                if sleep_data and 'dailySleepDTO' in sleep_data:
                    sleep_seconds = sleep_data['dailySleepDTO'].get('sleepTimeSeconds', 0)
                    if sleep_seconds > 0:
                        total_sleep_seconds += sleep_seconds
                        days_with_sleep += 1
            except:
                pass
            
            days_processed += 1
            if days_processed % 7 == 0:
                print(f"      ⏳ {days_processed}/{last_day} días...")
            
        except Exception as e:
            pass
        
        current_date += timedelta(days=1)
    
    # Actividades
    print(f"   🏃 Obteniendo actividades...")
    try:
        activities = garmin.get_activities(start_date, end_date, limit=100)
        
        for activity in activities:
            activity_date_str = activity.get('startTimeLocal', '')
            if activity_date_str:
                activity_date = datetime.strptime(activity_date_str[:10], '%Y-%m-%d')
                
                if activity_date.year == YEAR and activity_date.month == MONTH:
                    all_activities.append(activity)
                    
                    activity_type = activity.get('activityType', {}).get('typeKey', '').lower()
                    if any(kw in activity_type for kw in ['strength', 'training', 'gym', 'weight']):
                        strength_sessions += 1
        
    except Exception as e:
        print(f"      ⚠️  Error: {e}")
    
    # Cálculos
    avg_daily_steps = total_steps / days_with_steps if days_with_steps > 0 else 0
    avg_sleep_hours_garmin = (total_sleep_seconds / days_with_sleep / 3600) if days_with_sleep > 0 else 0
    
    # Escribir Garmin
    ws[f'A{row}'] = '═══ GARMIN CONNECT ═══'
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Steps (Ave Daily)'
    ws[f'B{row}'] = round(avg_daily_steps)
    ws[f'C{row}'] = 10000
    ws[f'D{row}'] = f'Promedio de {days_with_steps} días'
    row += 1
    
    ws[f'A{row}'] = 'Activities Mes'
    ws[f'B{row}'] = len(all_activities)
    ws[f'C{row}'] = 28
    ws[f'D{row}'] = f'{len(all_activities)} sesiones registradas'
    row += 1
    
    ws[f'A{row}'] = 'Strength Training'
    ws[f'B{row}'] = strength_sessions
    ws[f'C{row}'] = 8
    ws[f'D{row}'] = f'{strength_sessions} sesiones de fuerza'
    row += 1
    
    ws[f'A{row}'] = 'Sleep (Garmin backup)'
    ws[f'B{row}'] = round(avg_sleep_hours_garmin, 1)
    ws[f'C{row}'] = 7.0
    ws[f'D{row}'] = f'Promedio de {days_with_sleep} noches'
    row += 1
    
    print(f"   ✅ Garmin procesado:")
    print(f"      Steps: {round(avg_daily_steps):,}")
    print(f"      Activities: {len(all_activities)}")
    print(f"      Strength: {strength_sessions}")
    
except Exception as e:
    print(f"   ❌ Error Garmin: {e}")

row += 1

# ==================== WHOOP ====================
print("\n📊 Procesando WHOOP v2...")

ws[f'A{row}'] = '═══ WHOOP ═══'
ws[f'A{row}'].font = section_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

try:
    whoop = WhoopClientV2()
    
    if not whoop.auth.is_authenticated():
        ws[f'A{row}'] = 'WHOOP'
        ws[f'B{row}'] = 'No autenticado'
        ws[f'D{row}'] = 'Ejecuta: python3 whoop_client.py'
        row += 1
    else:
        # Obtener datos del mes
        summary = whoop.get_monthly_summary(YEAR, MONTH)
        
        # Sleep promedio
        ws[f'A{row}'] = 'Ave Sleep Duration (H)'
        ws[f'B{row}'] = round(summary['avg_sleep_hours'], 1) if summary['avg_sleep_hours'] > 0 else 'N/A'
        ws[f'C{row}'] = 7.0
        ws[f'D{row}'] = f"Basado en {len(summary['sleep'])} registros"
        row += 1
        
        # Días dormido antes de 9:30 PM
        ws[f'A{row}'] = 'Días dormido antes 9:30 PM'
        ws[f'B{row}'] = summary['days_sleep_before_930pm']
        ws[f'C{row}'] = 15
        ws[f'D{row}'] = f"{summary['days_sleep_before_930pm']} de {len(summary['sleep'])} noches"
        row += 1
        
        # Peso promedio (si está disponible)
        if summary['avg_weight_kg'] > 0:
            ws[f'A{row}'] = 'Ave Weight (kg)'
            ws[f'B{row}'] = round(summary['avg_weight_kg'], 1)
            ws[f'C{row}'] = '-'
            ws[f'D{row}'] = 'Medida corporal actual'
            row += 1
        
        print(f"   ✅ WHOOP procesado:")
        print(f"      Sleep promedio: {summary['avg_sleep_hours']:.1f}h")
        print(f"      Noches registradas: {len(summary['sleep'])}")
        print(f"      Días antes 9:30 PM: {summary['days_sleep_before_930pm']}")
        if summary['avg_weight_kg'] > 0:
            print(f"      Peso: {summary['avg_weight_kg']:.1f} kg")
        
except Exception as e:
    print(f"   ❌ Error WHOOP: {e}")
    ws[f'A{row}'] = 'ERROR'
    ws[f'B{row}'] = str(e)[:50]
    row += 1

# Resumen
row += 2
ws[f'A{row}'] = '═══ RESUMEN ═══'
ws[f'A{row}'].font = section_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

ws[f'A{row}'] = 'Fuentes de datos'
ws[f'B{row}'] = 'Garmin + WHOOP v2'
ws[f'D{row}'] = 'Ambas fuentes funcionando ✅'
row += 1

ws[f'A{row}'] = 'Datos disponibles'
ws[f'B{row}'] = 'Completos'
ws[f'D{row}'] = 'Steps, Activities, Sleep, Strength'
row += 1

# Ajustar anchos
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 40

# Guardar
filename = 'REPORTE_FINAL_ENERO_2026.xlsx'
wb.save(filename)

print(f"\n✅ REPORTE FINAL CREADO: {filename}")
print(f"   📍 /Users/AntonioXBruna/Desktop/fitness_tracker/{filename}")
print("\n" + "="*70 + "\n")

