"""
Crear Excel de prueba con datos de WHOOP y Garmin
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from garmin_client import GarminClient

print("\n" + "="*60)
print("CREANDO EXCEL DE PRUEBA")
print("="*60)

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Datos de Prueba"

# Título
ws['A1'] = 'REPORTE DE DATOS - FITNESS TRACKER'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws.merge_cells('A1:D1')

# Fecha
ws['A2'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws['A2'].font = Font(italic=True)

# Headers
ws['A4'] = 'Fuente'
ws['B4'] = 'Métrica'
ws['C4'] = 'Valor'
ws['D4'] = 'Notas'

for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}4'].font = Font(bold=True)
    ws[f'{col}4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Obtener datos de Garmin
print("\n📊 Obteniendo datos de Garmin...")
row = 5

try:
    garmin = GarminClient()
    garmin.login()
    
    stats = garmin.get_stats_for_date()
    
    if stats:
        # Pasos
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'Pasos'
        ws[f'C{row}'] = stats.get('totalSteps', 0)
        ws[f'D{row}'] = 'OK' if stats.get('totalSteps', 0) > 0 else 'Sin datos'
        row += 1
        
        # Calorías activas
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'Calorías Activas'
        ws[f'C{row}'] = stats.get('activeKilocalories', 0)
        ws[f'D{row}'] = 'OK' if stats.get('activeKilocalories', 0) > 0 else 'Bajo (normal si es temprano)'
        row += 1
        
        # Distancia
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'Distancia (km)'
        ws[f'C{row}'] = round(stats.get('totalDistanceMeters', 0) / 1000, 2)
        ws[f'D{row}'] = 'OK'
        row += 1
        
        # Calorías totales
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'Calorías Totales'
        ws[f'C{row}'] = stats.get('totalKilocalories', 0)
        ws[f'D{row}'] = 'Incluye BMR + activas'
        row += 1
        
        # Minutos de actividad
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'Minutos Activos'
        ws[f'C{row}'] = stats.get('activeTimeInSeconds', 0) / 60 if stats.get('activeTimeInSeconds') else 0
        ws[f'D{row}'] = 'En minutos'
        row += 1
        
        # Frecuencia cardíaca promedio
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'FC Promedio'
        ws[f'C{row}'] = stats.get('averageHeartRateInBeatsPerMinute', 'N/A')
        ws[f'D{row}'] = 'Promedio del día'
        row += 1
        
        # Frecuencia cardíaca en reposo
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'FC Reposo'
        ws[f'C{row}'] = stats.get('restingHeartRateInBeatsPerMinute', 'N/A')
        ws[f'D{row}'] = 'OK'
        row += 1
        
        print("   ✅ Datos de Garmin obtenidos")
        
    else:
        ws[f'A{row}'] = 'Garmin'
        ws[f'B{row}'] = 'ERROR'
        ws[f'C{row}'] = 'No se pudieron obtener datos'
        ws[f'D{row}'] = 'Verificar conexión'
        row += 1
        
except Exception as e:
    print(f"   ❌ Error con Garmin: {e}")
    ws[f'A{row}'] = 'Garmin'
    ws[f'B{row}'] = 'ERROR'
    ws[f'C{row}'] = str(e)
    ws[f'D{row}'] = 'Ver error en terminal'
    row += 1

# Espacio
row += 1

# Datos de WHOOP
print("\n📊 Obteniendo datos de WHOOP...")

try:
    from whoop_client import WhoopClient
    
    whoop = WhoopClient()
    
    # Perfil
    profile = whoop.get_profile()
    ws[f'A{row}'] = 'WHOOP'
    ws[f'B{row}'] = 'Usuario'
    ws[f'C{row}'] = f"{profile.get('first_name', '')} {profile.get('last_name', '')}"
    ws[f'D{row}'] = 'Autenticado ✅'
    row += 1
    
    ws[f'A{row}'] = 'WHOOP'
    ws[f'B{row}'] = 'User ID'
    ws[f'C{row}'] = profile.get('user_id', 'N/A')
    ws[f'D{row}'] = 'OK'
    row += 1
    
    # Nota sobre endpoints
    ws[f'A{row}'] = 'WHOOP'
    ws[f'B{row}'] = 'Recovery/Sleep'
    ws[f'C{row}'] = 'Error 404'
    ws[f'D{row}'] = 'Endpoints no disponibles - API cambió'
    row += 1
    
    print("   ✅ Perfil de WHOOP obtenido")
    print("   ⚠️  Recovery/Sleep endpoints dan 404")
    
except Exception as e:
    print(f"   ❌ Error con WHOOP: {e}")
    ws[f'A{row}'] = 'WHOOP'
    ws[f'B{row}'] = 'ERROR'
    ws[f'C{row}'] = str(e)
    ws[f'D{row}'] = 'Ver error'
    row += 1

# Espacio para resumen
row += 2
ws[f'A{row}'] = 'RESUMEN'
ws[f'A{row}'].font = Font(bold=True, size=12)
ws.merge_cells(f'A{row}:D{row}')
row += 1

ws[f'A{row}'] = '✅ Funcionando:'
ws[f'B{row}'] = 'Garmin Connect (pasos, calorías, distancia, FC)'
row += 1

ws[f'A{row}'] = '⚠️ Parcial:'
ws[f'B{row}'] = 'WHOOP (autenticación OK, pero endpoints 404)'
row += 1

ws[f'A{row}'] = '📊 Recomendación:'
ws[f'B{row}'] = 'Usar principalmente Garmin para datos actuales'
row += 1

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35

# Guardar
filename = f'test_datos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(filename)

print(f"\n✅ Excel de prueba creado: {filename}")
print(f"   Ubicación: /Users/AntonioXBruna/Desktop/fitness_tracker/{filename}")
print("\n" + "="*60 + "\n")

