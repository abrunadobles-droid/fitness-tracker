"""
Reporte formato final - Métricas en filas, meses en columnas
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from garmin_client import GarminClient
from whoop_client_v2 import WhoopClientV2
from calendar import monthrange

def get_monthly_data(year, month):
    """Obtiene datos de un mes específico"""
    print(f"\n📊 Procesando {month}/{year}...")
    
    data = {
        'steps_avg': 0,
        'activities': 0,
        'strength': 0,
        'sleep_hours': 0,
        'sleep_performance': 0,
        'sleep_consistency': 0,
        'days_before_930': 0,
        'recovery_score': 0,
        'hrv': 0,
        'hr_zone_1_3_monthly': 0,
        'hr_zone_4_5_monthly': 0
    }
    
    # GARMIN
    try:
        garmin = GarminClient()
        garmin.login()
        
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)
        
        total_steps = 0
        days_with_steps = 0
        all_activities = []
        strength_sessions = 0
        
        from datetime import timedelta
        current_date = start_date
        
        while current_date <= end_date:
            try:
                stats = garmin.get_stats_for_date(current_date)
                if stats and stats.get('totalSteps'):
                    total_steps += stats['totalSteps']
                    days_with_steps += 1
            except:
                pass
            current_date += timedelta(days=1)
        
        # Actividades
        activities = garmin.get_activities(start_date, end_date, limit=100)
        for activity in activities:
            activity_date_str = activity.get('startTimeLocal', '')
            if activity_date_str:
                activity_date = datetime.strptime(activity_date_str[:10], '%Y-%m-%d')
                if activity_date.year == year and activity_date.month == month:
                    all_activities.append(activity)
                    activity_type = activity.get('activityType', {}).get('typeKey', '').lower()
                    if any(kw in activity_type for kw in ['strength', 'training', 'gym', 'weight']):
                        strength_sessions += 1
        
        data['steps_avg'] = round(total_steps / days_with_steps) if days_with_steps > 0 else 0
        data['activities'] = len(all_activities)
        data['strength'] = strength_sessions
        
        print(f"   ✅ Garmin: {data['steps_avg']:,} steps, {data['activities']} activities, {data['strength']} strength")
        
    except Exception as e:
        print(f"   ⚠️  Garmin: {e}")
    
    # WHOOP
    try:
        whoop = WhoopClientV2()
        summary = whoop.get_monthly_summary(year, month)
        
        data['sleep_hours'] = round(summary['avg_sleep_hours'], 1)
        data['sleep_performance'] = round(summary['avg_sleep_performance'], 1)
        data['sleep_consistency'] = round(summary['avg_sleep_consistency'], 1)
        data['days_before_930'] = summary['days_sleep_before_930pm']
        data['recovery_score'] = round(summary['avg_recovery_score'], 1)
        data['hrv'] = round(summary['avg_hrv'], 1)
        
        # HR Zones convertidos a MENSUAL (4.3 semanas)
        data['hr_zone_1_3_monthly'] = round(summary['avg_time_hr_zone_1_3'] * 4.3, 1)
        data['hr_zone_4_5_monthly'] = round(summary['avg_time_hr_zone_4_5'] * 4.3, 1)
        
        print(f"   ✅ WHOOP: {data['sleep_hours']}h sleep, {data['sleep_consistency']}% consistency")
        
    except Exception as e:
        print(f"   ⚠️  WHOOP: {e}")
    
    return data


# ==================== CREAR EXCEL ====================
print("\n" + "="*70)
print("GENERANDO REPORTE ANUAL 2026")
print("="*70)

wb = Workbook()
ws = wb.active
ws.title = "Fitness Tracker 2026"

# Estilos
title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(bold=True, size=16, color='FFFFFF')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
section_font = Font(bold=True, size=11)

# Título
ws['A1'] = 'FITNESS & WELLNESS TRACKER 2026'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:N1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Headers de columnas
ws['A3'] = 'MÉTRICA'
ws['B3'] = 'META 2026'
months = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']
for i, month in enumerate(months, start=3):
    ws.cell(3, i, month).font = header_font
    ws.cell(3, i).fill = header_fill
    ws.cell(3, i).alignment = Alignment(horizontal='center')

ws['A3'].font = header_font
ws['B3'].font = header_font
ws['A3'].fill = header_fill
ws['B3'].fill = header_fill

# Métricas
metrics = [
    ('═══ GARMIN ═══', '', True),
    ('Steps (Ave Daily)', 10000, False),
    ('Activities Mes', 28, False),
    ('Strength Training', 10, False),
    ('═══ WHOOP ═══', '', True),
    ('Ave Sleep Duration (H)', 7.5, False),
    ('Sleep Performance %', 85, False),
    ('Sleep Consistency %', 85, False),
    ('Días dormido antes 9:30 PM', 15, False),
    ('Ave Recovery Score', 70, False),
    ('Ave HRV (ms)', '-', False),
    ('Time HR Zones 1-3 (H mensual)', 19.4, False),  # 4.5h x 4.3 semanas
    ('Time HR Zones 4-5 (H mensual)', 2.9, False),   # 40min x 4.3 semanas / 60
]

row = 4
for metric, meta, is_section in metrics:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = meta if meta != '' else ''
    
    if is_section:
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:N{row}')
    
    row += 1

# Obtener datos de Enero
enero_data = get_monthly_data(2026, 1)

# Llenar datos de Enero (columna C = índice 3)
ws['C5'] = enero_data['steps_avg']
ws['C6'] = enero_data['activities']
ws['C7'] = enero_data['strength']
ws['C9'] = enero_data['sleep_hours']
ws['C10'] = enero_data['sleep_performance']
ws['C11'] = enero_data['sleep_consistency']
ws['C12'] = enero_data['days_before_930']
ws['C13'] = enero_data['recovery_score']
ws['C14'] = enero_data['hrv']
ws['C15'] = enero_data['hr_zone_1_3_monthly']
ws['C16'] = enero_data['hr_zone_4_5_monthly']

# Ajustar anchos
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 12
for col in 'CDEFGHIJKLMN':
    ws.column_dimensions[col].width = 10

# Guardar
filename = 'FITNESS_TRACKER_2026.xlsx'
wb.save(filename)

print(f"\n✅ REPORTE CREADO: {filename}")
print("="*70 + "\n")

