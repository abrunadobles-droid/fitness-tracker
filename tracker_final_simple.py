"""
Fitness Tracker 2026 - Versión Final Simple
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from garmin_client import GarminClient
from whoop_client_v2 import WhoopClientV2
from calendar import monthrange

def get_monthly_data(year, month):
    print(f"\n📊 Procesando {month}/{year}...")
    
    data = {
        'steps_avg': 0, 'activities': 0, 'strength': 0,
        'days_before_930': 0, 'sleep_hours': 0,
        'hr_zone_1_3_monthly': 0, 'hr_zone_4_5_monthly': 0,
        'sleep_performance': 0, 'sleep_consistency': 0,
        'recovery_score': 0, 'hrv': 0
    }
    
    try:
        garmin = GarminClient()
        garmin.login()
        
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)
        
        total_steps = 0
        days_with_steps = 0
        all_activities = []
        strength_sessions = 0
        
        current_date = start_date
        while current_date <= end_date:
            try:
                stats = garmin.get_stats_for_date(current_date)
                if stats and stats.get('totalSteps'):
                    total_steps += stats['totalSteps']
                    days_with_steps += 1
            except:
                pass
            current_date += timedelta(days=1)
        
        activities = garmin.get_activities(start_date, end_date, limit=100)
        for activity in activities:
            activity_date_str = activity.get('startTimeLocal', '')
            if activity_date_str:
                activity_date = datetime.strptime(activity_date_str[:10], '%Y-%m-%d')
                if activity_date.year == year and activity_date.month == month:
                    all_activities.append(activity)
                    activity_type = activity.get('activityType', {}).get('typeKey', '').lower()
                    if any(kw in activity_type for kw in ['strength', 'training', 'gym', 'weight']):
                        strength_sessions += 1
        
        data['steps_avg'] = round(total_steps / days_with_steps) if days_with_steps > 0 else 0
        data['activities'] = len(all_activities)
        data['strength'] = strength_sessions
        
    except Exception as e:
        print(f"   ⚠️  Garmin: {e}")
    
    try:
        whoop = WhoopClientV2()
        summary = whoop.get_monthly_summary(year, month)
        
        data['days_before_930'] = summary['days_sleep_before_930pm']
        data['sleep_hours'] = round(summary['avg_sleep_hours'], 1)
        data['hr_zone_1_3_monthly'] = round(summary['avg_time_hr_zone_1_3'] * 4.3, 1)
        data['hr_zone_4_5_monthly'] = round(summary['avg_time_hr_zone_4_5'] * 4.3, 1)
        data['sleep_performance'] = round(summary['avg_sleep_performance'], 1)
        data['sleep_consistency'] = round(summary['avg_sleep_consistency'], 1)
        data['recovery_score'] = round(summary['avg_recovery_score'], 1)
        data['hrv'] = round(summary['avg_hrv'], 1)
        
    except Exception as e:
        print(f"   ⚠️  WHOOP: {e}")
    
    return data

print("\n" + "="*70)
print("FITNESS & WELLNESS TRACKER 2026")
print("="*70)

wb = Workbook()
ws = wb.active
ws.title = "Tracker 2026"

title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(bold=True, size=16, color='FFFFFF')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
resultado_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

# Título
ws['A1'] = 'FITNESS & WELLNESS TRACKER 2026'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:N1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Headers
ws['A3'] = 'MÉTRICA'
ws['B3'] = 'META'
months = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']
for i, month in enumerate(months, start=3):
    cell = ws.cell(3, i, month)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

ws['A3'].font = header_font
ws['B3'].font = header_font
ws['A3'].fill = header_fill
ws['B3'].fill = header_fill

# HÁBITOS
row = 4
ws[f'A{row}'] = '🎯 HÁBITOS'
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:N{row}')
row += 1

habitos_start_row = row
ws[f'A{row}'] = 'Steps (Ave Daily)'
ws[f'B{row}'] = 10000
row += 1
ws[f'A{row}'] = 'Activities Mes'
ws[f'B{row}'] = 28
row += 1
ws[f'A{row}'] = 'Strength Training'
ws[f'B{row}'] = 10
row += 1
ws[f'A{row}'] = 'Días dormido antes 9:30 PM'
ws[f'B{row}'] = 15
row += 1
ws[f'A{row}'] = 'Ave Sleep Duration (H)'
ws[f'B{row}'] = 7.5
row += 1
ws[f'A{row}'] = 'Time HR Zones 1-3 (H mensual)'
ws[f'B{row}'] = 19.4
row += 1
ws[f'A{row}'] = 'Time HR Zones 4-5 (H mensual)'
ws[f'B{row}'] = 2.9
habitos_end_row = row
row += 1

# Espacio
row += 1

# RESULTADOS
ws[f'A{row}'] = '📊 RESULTADOS'
ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
ws[f'A{row}'].fill = resultado_fill
ws.merge_cells(f'A{row}:N{row}')
row += 1

resultados_start_row = row
ws[f'A{row}'] = 'Sleep Performance %'
ws[f'B{row}'] = 85
row += 1
ws[f'A{row}'] = 'Sleep Consistency %'
ws[f'B{row}'] = 85
row += 1
ws[f'A{row}'] = 'Ave Recovery Score'
ws[f'B{row}'] = 70
row += 1
ws[f'A{row}'] = 'Ave HRV (ms)'
ws[f'B{row}'] = '-'
row += 1

# Obtener datos
enero_data = get_monthly_data(2026, 1)

# Llenar Enero en HÁBITOS
ws['C5'] = enero_data['steps_avg']
ws['C6'] = enero_data['activities']
ws['C7'] = enero_data['strength']
ws['C8'] = enero_data['days_before_930']
ws['C9'] = enero_data['sleep_hours']
ws['C10'] = enero_data['hr_zone_1_3_monthly']
ws['C11'] = enero_data['hr_zone_4_5_monthly']

# Llenar Enero en RESULTADOS
ws['C14'] = enero_data['sleep_performance']
ws['C15'] = enero_data['sleep_consistency']
ws['C16'] = enero_data['recovery_score']
ws['C17'] = enero_data['hrv']

# Formato condicional
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for col_idx in range(3, 15):
    col_letter = chr(64 + col_idx)
    for row_idx in range(habitos_start_row, habitos_end_row + 1):
        formula_cell = f'$B${row_idx}'
        value_cell = f'{col_letter}{row_idx}'
        
        ws.conditional_formatting.add(
            f'{value_cell}:{value_cell}',
            CellIsRule(operator='greaterThanOrEqual', formula=[formula_cell], fill=green_fill)
        )
        ws.conditional_formatting.add(
            f'{value_cell}:{value_cell}',
            CellIsRule(operator='lessThan', formula=[formula_cell], fill=red_fill)
        )

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 10
for col in 'CDEFGHIJKLMN':
    ws.column_dimensions[col].width = 10

wb.save('FITNESS_TRACKER_2026.xlsx')

print(f"\n✅ TRACKER CREADO: FITNESS_TRACKER_2026.xlsx")
print(f"   📊 Enero completado con formato verde/rojo")
print("="*70 + "\n")

