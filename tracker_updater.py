"""
Módulo principal para actualizar el Excel Tracker
Integra datos de WHOOP y Garmin Connect
"""

import openpyxl
from datetime import datetime
from calendar import monthrange
from whoop_client import WhoopClient
from garmin_client import GarminClient
from config import EXCEL_FILE, SHEET_NAME


class FitnessTrackerUpdater:
    def __init__(self, excel_path=None):
        self.excel_path = excel_path or EXCEL_FILE
        self.whoop = WhoopClient()
        self.garmin = GarminClient()
        
        # Mapeo de métricas a filas del Excel
        # Estos números de fila son aproximados, hay que ajustarlos según tu Excel real
        self.metric_rows = {
            'Activities Mes': 3,
            'Active Calories': 4,
            'Steps (Ave Daily)': 5,
            'Strenght Training': 6,
            'Ave Sleep Duration (H)': 7,
        }
        
        # Mapeo de columnas para meses (año 2026)
        # Columnas basadas en estructura vista: Enero 2026 está en columna 23 (index 22)
        self.month_columns = {
            1: 23,   # Enero
            2: 24,   # Febrero
            3: 25,   # Marzo
            4: 26,   # Abril
            5: 27,   # Mayo
            6: 28,   # Junio
            7: 29,   # Julio
            8: 30,   # Agosto
            9: 31,   # Setiembre
            10: 32,  # Octubre
            11: 33,  # Noviembre
            12: 34,  # Diciembre
        }
    
    def load_workbook(self):
        """Carga el workbook de Excel"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            return wb
        except FileNotFoundError:
            print(f"❌ No se encontró el archivo: {self.excel_path}")
            print("   Asegúrate de que la ruta en config.py es correcta")
            return None
    
    def update_monthly_data(self, year, month, dry_run=True):
        """
        Actualiza datos mensuales en el Excel
        
        Args:
            year: Año (ej: 2026)
            month: Mes (1-12)
            dry_run: Si es True, solo muestra qué se actualizaría sin guardar
        """
        print(f"\n{'='*60}")
        print(f"ACTUALIZANDO DATOS: {month}/{year}")
        print(f"{'='*60}\n")
        
        # Obtener datos de las APIs
        print("📊 Obteniendo datos de WHOOP...")
        try:
            whoop_data = self.whoop.get_monthly_summary(year, month)
            print(f"   ✅ Datos WHOOP obtenidos")
        except Exception as e:
            print(f"   ⚠️  Error con WHOOP: {e}")
            whoop_data = {}
        
        print("\n📊 Obteniendo datos de Garmin...")
        try:
            self.garmin.login()
            garmin_data = self.garmin.get_monthly_summary(year, month)
            print(f"   ✅ Datos Garmin obtenidos")
        except Exception as e:
            print(f"   ⚠️  Error con Garmin: {e}")
            garmin_data = {}
        
        # Preparar datos para actualizar
        updates = {}
        
        # Steps (de Garmin)
        if garmin_data.get('avg_daily_steps'):
            updates['Steps (Ave Daily)'] = round(garmin_data['avg_daily_steps'])
        
        # Active Calories (de Garmin)
        if garmin_data.get('total_active_calories'):
            updates['Active Calories'] = round(garmin_data['total_active_calories'])
        
        # Sleep Duration (preferir WHOOP, fallback a Garmin)
        sleep_hours = whoop_data.get('avg_sleep_hours') or garmin_data.get('avg_sleep_hours')
        if sleep_hours:
            updates['Ave Sleep Duration (H)'] = round(sleep_hours, 1)
        
        # Strength Training (de Garmin)
        if garmin_data.get('strength_training_sessions'):
            updates['Strenght Training'] = garmin_data['strength_training_sessions']
        
        # Activities (workouts de WHOOP)
        if whoop_data.get('total_workouts'):
            updates['Activities Mes'] = whoop_data['total_workouts']
        
        # Mostrar qué se va a actualizar
        print(f"\n📝 Datos a actualizar para {month}/{year}:")
        print("-" * 60)
        for metric, value in updates.items():
            print(f"   {metric:30} → {value}")
        
        if dry_run:
            print(f"\n⚠️  Modo DRY RUN - No se guardaron cambios")
            print("   Para guardar cambios, ejecuta con dry_run=False")
            return updates
        
        # Actualizar Excel
        print(f"\n💾 Actualizando Excel...")
        wb = self.load_workbook()
        if not wb:
            return None
        
        try:
            sheet = wb[SHEET_NAME]
            
            # Obtener columna para este mes
            col = self.month_columns.get(month)
            if not col:
                print(f"❌ No hay columna configurada para el mes {month}")
                return None
            
            # Actualizar cada métrica
            for metric, value in updates.items():
                row = self.metric_rows.get(metric)
                if row:
                    # +1 porque Excel es 1-indexed
                    cell = sheet.cell(row=row, column=col)
                    cell.value = value
                    print(f"   ✅ Actualizado {metric} en celda {cell.coordinate}")
            
            # Guardar
            wb.save(self.excel_path)
            print(f"\n✅ Excel actualizado exitosamente!")
            print(f"   Archivo: {self.excel_path}")
            
            return updates
            
        except Exception as e:
            print(f"\n❌ Error actualizando Excel: {e}")
            return None
    
    def update_current_month(self, dry_run=True):
        """Actualiza el mes actual"""
        now = datetime.now()
        return self.update_monthly_data(now.year, now.month, dry_run)
    
    def get_current_stats(self):
        """Obtiene estadísticas del día actual de todas las fuentes"""
        print(f"\n{'='*60}")
        print(f"ESTADÍSTICAS DEL DÍA - {datetime.now().strftime('%d/%m/%Y')}")
        print(f"{'='*60}\n")
        
        # WHOOP
        print("📱 WHOOP:")
        try:
            if not self.whoop.auth.is_authenticated():
                print("   ⚠️  No autenticado - ejecuta whoop_client.py primero")
            else:
                recovery = self.whoop.get_recovery_for_date()
                if recovery:
                    rec = recovery[0]['score']
                    print(f"   Recovery Score: {rec['recovery_score']}")
                    print(f"   HRV: {rec['hrv_rmssd_milli']} ms")
                    print(f"   Resting HR: {rec['resting_heart_rate']} bpm")
                else:
                    print("   No hay datos de recovery")
                
                sleep = self.whoop.get_sleep_for_date()
                if sleep:
                    slp = sleep[0]['score']
                    hours = slp['stage_summary']['total_in_bed_time_milli'] / 3600000
                    print(f"   Sleep: {hours:.1f}h")
        except Exception as e:
            print(f"   Error: {e}")
        
        # Garmin
        print("\n⌚ GARMIN:")
        try:
            self.garmin.login()
            stats = self.garmin.get_stats_for_date()
            if stats:
                print(f"   Steps: {stats.get('totalSteps', 0):,}")
                print(f"   Active Calories: {stats.get('activeKilocalories', 0):,}")
                print(f"   Distance: {stats.get('totalDistanceMeters', 0)/1000:.2f} km")
            
            sleep_hours = self.garmin.get_sleep_duration_hours()
            if sleep_hours > 0:
                print(f"   Sleep: {sleep_hours:.1f}h")
        except Exception as e:
            print(f"   Error: {e}")
        
        print(f"\n{'='*60}\n")


def main():
    """Función principal de ejemplo"""
    import sys
    
    tracker = FitnessTrackerUpdater()
    
    # Si se pasa "update" como argumento, actualiza el mes actual
    if len(sys.argv) > 1 and sys.argv[1] == "update":
        # Dry run primero
        print("\n🔍 Ejecutando DRY RUN...")
        tracker.update_current_month(dry_run=True)
        
        # Pedir confirmación
        print("\n¿Deseas guardar estos cambios? (s/n): ", end='')
        if input().lower() == 's':
            tracker.update_current_month(dry_run=False)
        else:
            print("Operación cancelada")
    else:
        # Por defecto, solo mostrar stats actuales
        tracker.get_current_stats()


if __name__ == '__main__':
    main()
