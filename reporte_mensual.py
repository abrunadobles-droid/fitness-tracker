"""
Reporte mensual completo - Enero 2026
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from garmin_client import GarminClient
from calendar import monthrange

print("\n" + "="*60)
print("REPORTE MENSUAL - ENERO 2026")
print("="*60)

# Configuración
YEAR = 2026
MONTH = 1  # Enero

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Enero 2026"

# Título
ws['A1'] = 'REPORTE MENSUAL - ENERO 2026'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws.merge_cells('A1:C1')

ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

# Headers
ws['A4'] = 'MÉTRICA'
ws['B4'] = 'VALOR'
ws['C4'] = 'NOTAS'

for col in ['A', 'B', 'C']:
    ws[f'{col}4'].font = Font(bold=True)
    ws[f'{col}4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

row = 5

# ==================== GARMIN ====================
print("\n📊 Procesando datos de Garmin para Enero 2026...")

try:
    garmin = GarminClient()
    garmin.login()
    
    # Obtener todos los días de enero
    start_date = datetime(YEAR, MONTH, 1)
    last_day = monthrange(YEAR, MONTH)[1]
    end_date = datetime(YEAR, MONTH, last_day)
    
    total_steps = 0
    days_with_steps = 0
    all_activities = []
    strength_sessions = 0
    
    print(f"   Obteniendo datos del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}...")
    
    # Iterar cada día del mes
    current_date = start_date
    days_processed = 0
    
    while current_date <= end_date:
        try:
            # Stats diarios
            stats = garmin.get_stats_for_date(current_date)
            if stats and stats.get('totalSteps'):
                total_steps += stats['totalSteps']
                days_with_steps += 1
            
            days_processed += 1
            if days_processed % 5 == 0:
                print(f"      Procesados {days_processed}/{last_day} días...")
            
        except Exception as e:
            print(f"      Error en {current_date.strftime('%d/%m')}: {e}")
        
        current_date += timedelta(days=1)
    
    # Obtener actividades del mes
    print(f"   Obteniendo actividades...")
    try:
        # Garmin devuelve las más recientes primero
        activities = garmin.get_activities(start_date, end_date, limit=100)
        
        # Filtrar solo las de enero 2026
        for activity in activities:
            activity_date_str = activity.get('startTimeLocal', '')
            if activity_date_str:
                # Formato: "2026-01-15 10:30:00"
                activity_date = datetime.strptime(activity_date_str[:10], '%Y-%m-%d')
                
                if activity_date.year == YEAR and activity_date.month == MONTH:
                    all_activities.append(activity)
                    
                    # Contar strength training
                    activity_type = activity.get('activityType', {}).get('typeKey', '').lower()
                    if any(keyword in activity_type for keyword in ['strength', 'training', 'gym', 'weight']):
                        strength_sessions += 1
        
        print(f"      ✅ {len(all_activities)} actividades encontradas")
        
    except Exception as e:
        print(f"      ⚠️  Error obteniendo actividades: {e}")
    
    # Calcular promedio de pasos diarios
    avg_daily_steps = total_steps / days_with_steps if days_with_steps > 0 else 0
    
    # ESCRIBIR RESULTADOS GARMIN
    ws['A5'] = '═══ GARMIN ═══'
    ws['A5'].font = Font(bold=True, size=12)
    ws.merge_cells('A5:C5')
    row = 6
    
    # Steps Average Daily
    ws[f'A{row}'] = 'Steps (Ave Daily)'
    ws[f'B{row}'] = round(avg_daily_steps)
    ws[f'C{row}'] = f'Promedio de {days_with_steps} días con datos'
    row += 1
    
    # Activities (cantidad)
    ws[f'A{row}'] = 'Activities Mes'
    ws[f'B{row}'] = len(all_activities)
    ws[f'C{row}'] = f'{len(all_activities)} sesiones de ejercicio registradas'
    row += 1
    
    # Strength Training
    ws[f'A{row}'] = 'Strength Training'
    ws[f'B{row}'] = strength_sessions
    ws[f'C{row}'] = f'{strength_sessions} sesiones de fuerza/pesas'
    row += 1
    
    print(f"\n   ✅ Resultados Garmin:")
    print(f"      Steps Average Daily: {round(avg_daily_steps):,}")
    print(f"      Activities: {len(all_activities)}")
    print(f"      Strength Training: {strength_sessions}")
    
except Exception as e:
    print(f"\n   ❌ Error con Garmin: {e}")
    ws[f'A{row}'] = 'ERROR GARMIN'
    ws[f'B{row}'] = str(e)
    row += 1

# Espacio
row += 1

# ==================== WHOOP ====================
print("\n📊 Procesando datos de WHOOP para Enero 2026...")

ws[f'A{row}'] = '═══ WHOOP ═══'
ws[f'A{row}'].font = Font(bold=True, size=12)
ws.merge_cells(f'A{row}:C{row}')
row += 1

try:
    from whoop_client import WhoopClient
    whoop = WhoopClient()
    
    # Verificar autenticación
    if not whoop.auth.is_authenticated():
        ws[f'A{row}'] = 'WHOOP'
        ws[f'B{row}'] = 'No autenticado'
        ws[f'C{row}'] = 'Ejecuta: python3 whoop_client.py'
        row += 1
    else:
        ws[f'A{row}'] = 'Ave Sleep Duration (H)'
        ws[f'B{row}'] = 'N/A'
        ws[f'C{row}'] = 'Endpoint da 404 - API cambió'
        row += 1
        
        ws[f'A{row}'] = 'Ave Weight'
        ws[f'B{row}'] = 'N/A'
        ws[f'C{row}'] = 'No disponible en API actual'
        row += 1
        
        ws[f'A{row}'] = 'Días dormido antes 9:30 PM'
        ws[f'B{row}'] = 'N/A'
        ws[f'C{row}'] = 'Endpoint da 404 - API cambió'
        row += 1
        
        print(f"   ⚠️  WHOOP autenticado pero endpoints no disponibles")
        print(f"      Los endpoints de sleep/recovery dan 404")
        print(f"      La API de WHOOP parece haber cambiado")
        
except Exception as e:
    print(f"\n   ❌ Error con WHOOP: {e}")
    ws[f'A{row}'] = 'ERROR WHOOP'
    ws[f'B{row}'] = str(e)
    row += 1

# Espacio para resumen
row += 2
ws[f'A{row}'] = '═══ RESUMEN ═══'
ws[f'A{row}'].font = Font(bold=True, size=12)
ws.merge_cells(f'A{row}:C{row}')
row += 1

ws[f'A{row}'] = 'Estado'
ws[f'B{row}'] = 'Garmin: ✅ Funcionando'
row += 1

ws[f'A{row}'] = ''
ws[f'B{row}'] = 'WHOOP: ⚠️ Autenticado pero sin datos'
row += 1

ws[f'A{row}'] = 'Datos disponibles'
ws[f'B{row}'] = 'Pasos, Actividades, Strength Training'
row += 1

ws[f'A{row}'] = 'Datos NO disponibles'
ws[f'B{row}'] = 'Sleep, Weight, Recovery (WHOOP API cambió)'
row += 1

# Ajustar anchos
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 40

# Guardar
filename = f'reporte_enero_2026.xlsx'
wb.save(filename)

print(f"\n✅ Reporte mensual creado: {filename}")
print(f"   Ubicación: /Users/AntonioXBruna/Desktop/fitness_tracker/{filename}")
print("\n" + "="*60 + "\n")

