"""
Reporte Final Completo - Enero 2026
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from garmin_client import GarminClient
from whoop_client_v2 import WhoopClientV2
from calendar import monthrange

print("\n" + "="*70)
print("  REPORTE FINAL COMPLETO - ENERO 2026")
print("="*70)

YEAR = 2026
MONTH = 1

wb = Workbook()
ws = wb.active
ws.title = "Enero 2026"

# Estilos
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, size=14, color='FFFFFF')
subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
subheader_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=12, color='4472C4')

# Título
ws['A1'] = 'REPORTE MENSUAL FITNESS & WELLNESS'
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws.merge_cells('A1:D1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

ws['A2'] = f'Enero 2026'
ws['A2'].font = Font(bold=True, size=12)
ws.merge_cells('A2:D2')

ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws['A3'].font = Font(italic=True, size=9)

ws['A5'] = 'MÉTRICA'
ws['B5'] = 'VALOR'
ws['C5'] = 'META 2026'
ws['D5'] = 'NOTAS'

for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}5'].font = subheader_font
    ws[f'{col}5'].fill = subheader_fill
    ws[f'{col}5'].alignment = Alignment(horizontal='center')

row = 6

# GARMIN
print("\n📊 Procesando GARMIN...")

try:
    garmin = GarminClient()
    garmin.login()
    
    start_date = datetime(YEAR, MONTH, 1)
    last_day = monthrange(YEAR, MONTH)[1]
    end_date = datetime(YEAR, MONTH, last_day)
    
    total_steps = 0
    days_with_steps = 0
    all_activities = []
    strength_sessions = 0
    
    current_date = start_date
    days_processed = 0
    
    while current_date <= end_date:
        try:
            stats = garmin.get_stats_for_date(current_date)
            if stats and stats.get('totalSteps'):
                total_steps += stats['totalSteps']
                days_with_steps += 1
            
            days_processed += 1
            if days_processed % 7 == 0:
                print(f"      ⏳ {days_processed}/{last_day} días...")
        except:
            pass
        
        current_date += timedelta(days=1)
    
    # Actividades
    try:
        activities = garmin.get_activities(start_date, end_date, limit=100)
        
        for activity in activities:
            activity_date_str = activity.get('startTimeLocal', '')
            if activity_date_str:
                activity_date = datetime.strptime(activity_date_str[:10], '%Y-%m-%d')
                
                if activity_date.year == YEAR and activity_date.month == MONTH:
                    all_activities.append(activity)
                    
                    activity_type = activity.get('activityType', {}).get('typeKey', '').lower()
                    if any(kw in activity_type for kw in ['strength', 'training', 'gym', 'weight']):
                        strength_sessions += 1
    except:
        pass
    
    avg_daily_steps = total_steps / days_with_steps if days_with_steps > 0 else 0
    
    ws[f'A{row}'] = '═══ GARMIN CONNECT ═══'
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Steps (Ave Daily)'
    ws[f'B{row}'] = round(avg_daily_steps)
    ws[f'C{row}'] = 10000
    ws[f'D{row}'] = f'Promedio de {days_with_steps} días'
    row += 1
    
    ws[f'A{row}'] = 'Activities Mes'
    ws[f'B{row}'] = len(all_activities)
    ws[f'C{row}'] = 28
    ws[f'D{row}'] = f'{len(all_activities)} sesiones'
    row += 1
    
    ws[f'A{row}'] = 'Strength Training'
    ws[f'B{row}'] = strength_sessions
    ws[f'C{row}'] = 8
    ws[f'D{row}'] = f'{strength_sessions} sesiones de fuerza'
    row += 1
    
    print(f"   ✅ Steps: {round(avg_daily_steps):,}, Activities: {len(all_activities)}, Strength: {strength_sessions}")
    
except Exception as e:
    print(f"   ❌ Error: {e}")

row += 1

# WHOOP
print("\n📊 Procesando WHOOP...")

ws[f'A{row}'] = '═══ WHOOP ═══'
ws[f'A{row}'].font = section_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

try:
    whoop = WhoopClientV2()
    summary = whoop.get_monthly_summary(YEAR, MONTH)
    
    ws[f'A{row}'] = 'Ave Sleep Duration (H)'
    ws[f'B{row}'] = round(summary['avg_sleep_hours'], 1)
    ws[f'C{row}'] = 7.0
    ws[f'D{row}'] = f"Tiempo real dormido ({len(summary['sleep'])} noches)"
    row += 1
    
    ws[f'A{row}'] = 'Sleep Performance %'
    ws[f'B{row}'] = round(summary['avg_sleep_performance'], 1)
    ws[f'C{row}'] = 85
    ws[f'D{row}'] = 'Promedio mensual'
    row += 1
    
    ws[f'A{row}'] = 'Sleep Consistency %'
    ws[f'B{row}'] = round(summary['avg_sleep_consistency'], 1)
    ws[f'C{row}'] = 80
    ws[f'D{row}'] = 'Promedio mensual'
    row += 1
    
    ws[f'A{row}'] = 'Días dormido antes 9:30 PM'
    ws[f'B{row}'] = summary['days_sleep_before_930pm']
    ws[f'C{row}'] = 15
    ws[f'D{row}'] = f"{summary['days_sleep_before_930pm']}/{len(summary['sleep'])} noches"
    row += 1
    
    ws[f'A{row}'] = 'Ave Recovery Score'
    ws[f'B{row}'] = round(summary['avg_recovery_score'], 1)
    ws[f'C{row}'] = 70
    ws[f'D{row}'] = 'Promedio mensual'
    row += 1
    
    ws[f'A{row}'] = 'Ave HRV (ms)'
    ws[f'B{row}'] = round(summary['avg_hrv'], 1)
    ws[f'C{row}'] = '-'
    ws[f'D{row}'] = 'Heart Rate Variability'
    row += 1
    
    ws[f'A{row}'] = 'Time HR Zones 1-3 (min)'
    ws[f'B{row}'] = round(summary['avg_time_hr_zone_1_3'], 1)
    ws[f'C{row}'] = '-'
    ws[f'D{row}'] = f'Promedio por workout'
    row += 1
    
    ws[f'A{row}'] = 'Time HR Zones 4-5 (min)'
    ws[f'B{row}'] = round(summary['avg_time_hr_zone_4_5'], 1)
    ws[f'C{row}'] = '-'
    ws[f'D{row}'] = 'Alta intensidad'
    row += 1
    
    print(f"   ✅ Todas las métricas WHOOP procesadas")
    
except Exception as e:
    print(f"   ❌ Error: {e}")

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 45

filename = 'ENERO_2026_COMPLETO.xlsx'
wb.save(filename)

print(f"\n✅ REPORTE CREADO: {filename}")
print("="*70 + "\n")

